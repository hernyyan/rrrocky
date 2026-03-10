"""
Excel processing service.
Handles reading Excel workbooks and converting sheets to CSVs for Layer 1 extraction.
PDF conversion has been removed — the browser now renders the original file client-side.
"""
import csv
import io
import os
import re
import tempfile
import warnings
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Dict, List

import openpyxl

# Monkey-patch openpyxl to skip invalid defined names (e.g. corrupt print titles
# from QuickBooks, Sage, NetSuite exports). Without this, load_workbook crashes
# on files with entries like "Formula removed, name can be deleted."
from openpyxl.reader import workbook as _wb_reader
_orig_assign_names = _wb_reader.WorkbookParser.assign_names
def _safe_assign_names(self):
    try:
        _orig_assign_names(self)
    except Exception:
        pass  # Skip all name assignment errors — we don't need print titles or named ranges
_wb_reader.WorkbookParser.assign_names = _safe_assign_names


def _open_workbook(filepath: str, read_only: bool = False):
    """
    Open an Excel workbook with resilient error handling.

    Attempt 1: Normal load with monkey-patched assign_names (handles most cases).
    Attempt 2: Strip definedNames and calcChain from the ZIP and retry (handles the rest).

    Portfolio company Excel files from accounting software frequently contain
    invalid XML that crashes openpyxl's strict parser.
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            return openpyxl.load_workbook(filepath, read_only=read_only, data_only=True)
        except Exception as e:
            print(f"[excel_processor] Normal load failed: {e}")
            print(f"[excel_processor] Attempting ZIP-level repair...")

    # Attempt 2: Strip problematic XML and retry
    try:
        repaired_path = _repair_workbook(filepath)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(repaired_path, read_only=read_only, data_only=True)
        try:
            os.unlink(repaired_path)
        except OSError:
            pass
        return wb
    except Exception as e2:
        print(f"[excel_processor] Repair also failed: {e2}")
        raise ValueError(
            f"Unable to read workbook even after repair. Try opening the file in "
            f"Excel, saving as a new .xlsx, and re-uploading. Error: {e2}"
        )


def _repair_workbook(filepath: str) -> str:
    """
    Create a repaired copy of an xlsx by stripping problematic XML elements.
    Removes definedNames (print titles, named ranges) and calcChain.xml.
    """
    temp_dir = tempfile.mkdtemp()
    repaired_path = os.path.join(temp_dir, "repaired.xlsx")

    with zipfile.ZipFile(filepath, 'r') as zin:
        with zipfile.ZipFile(repaired_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)

                if item.filename == 'xl/workbook.xml':
                    try:
                        for prefix, uri in [
                            ('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'),
                            ('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'),
                        ]:
                            ET.register_namespace(prefix, uri)
                        root = ET.fromstring(data)
                        ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                        for elem in root.findall(f'{{{ns}}}definedNames'):
                            root.remove(elem)
                        data = ET.tostring(root, xml_declaration=True, encoding='UTF-8')
                        print("[excel_processor] Stripped definedNames from workbook.xml")
                    except ET.ParseError:
                        pass

                if item.filename == 'xl/calcChain.xml':
                    print("[excel_processor] Skipping calcChain.xml")
                    continue

                zout.writestr(item, data)

    return repaired_path


def _safe_filename(name: str) -> str:
    """Sanitize a sheet name for use as a filename."""
    safe = re.sub(r'[\\/*?:"<>|]', "_", name)
    safe = safe.strip(". ")
    return safe or "sheet"


def get_sheet_names(filepath: str) -> List[str]:
    """
    Return an ordered list of visible sheet names from the workbook.

    Args:
        filepath: Absolute path to the .xlsx/.xls file.

    Returns:
        List of visible sheet names in workbook order.
    """
    wb = _open_workbook(filepath, read_only=True)
    names = [
        name for name in wb.sheetnames
        if wb[name].sheet_state == "visible"
    ]
    wb.close()
    return names


def convert_to_csvs(filepath: str) -> Dict[str, str]:
    """
    Convert each visible sheet in an Excel workbook to CSV content using openpyxl.

    Args:
        filepath: Absolute path to the uploaded .xlsx / .xls file.

    Returns:
        Mapping of sheet_name → CSV content string.
        Content is NOT written to disk — the caller persists it.

    Notes:
        - Hidden and empty sheets are skipped.
        - All cell content is passed as-is; the AI prompt handles interpretation.
        - Merged cells: openpyxl returns the value in the top-left cell, None elsewhere.
    """
    source_path = Path(filepath)
    if not source_path.exists():
        raise FileNotFoundError(f"Upload file not found: {filepath}")

    wb = _open_workbook(filepath, read_only=False)
    csv_contents: Dict[str, str] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if ws.sheet_state != "visible":
            continue

        if ws.max_row is None or ws.max_row == 0:
            continue

        output = io.StringIO()
        writer = csv.writer(output)

        for row in ws.iter_rows(values_only=True):
            writer.writerow([("" if cell is None else str(cell)) for cell in row])

        csv_contents[sheet_name] = output.getvalue()

    wb.close()
    return csv_contents
