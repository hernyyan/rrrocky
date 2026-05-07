"""
Python-side Excel extraction for the Layer 1 4-step pipeline.

Step A: extract_header_rows  — reads the first N rows of a sheet and returns
         them as a plain-text block for the AI column identifier.

Step C: extract_rows_with_metadata — reads the full sheet for the identified
         data column, returning one dict per row with value, formatting flags
         (bold, italic, indent), and the row label.

         rows_to_csv_with_metadata — serialises that list to a CSV string that
         the AI structured extractor can consume.
"""
from __future__ import annotations

import csv
import io
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Font


# ── helpers ──────────────────────────────────────────────────────────────────

def _effective_font(cell) -> Font:
    """Return the cell's font, falling back to a plain Font() if absent."""
    return cell.font if cell.font else Font()


def _indent_level(cell) -> int:
    """Return the cell's alignment indent level (0 if not set)."""
    if cell.alignment and cell.alignment.indent:
        return int(cell.alignment.indent)
    return 0


def _cell_value(cell) -> Any:
    return cell.value


# ── Step A ────────────────────────────────────────────────────────────────────

def extract_header_rows(
    filepath: str,
    sheet_name: str,
    n_rows: int = 150,
) -> str:
    """
    Open the workbook and return the first *n_rows* rows of *sheet_name* as a
    plain-text block with 1-based row numbers prepended.

    Format: "[row_num]\tcol1\tcol2\t..."

    Row numbers allow the AI column identifier (Step B) to return precise
    section_start_row / section_end_row values for multi-statement sheets.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    lines: List[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= n_rows:
            break
        row_num = i + 1
        cells = [str(v) if v is not None else "" for v in row]
        lines.append(f"[{row_num}]\t" + "\t".join(cells))

    wb.close()
    return "\n".join(lines)


# ── Step C ────────────────────────────────────────────────────────────────────

def extract_rows_with_metadata(
    filepath: str,
    sheet_name: str,
    column_index: int,          # 1-based column index of the target period
    source_scaling: str,        # 'thousands' | 'millions' | 'actual_dollars'
    skip_rows: int = 0,         # legacy: rows to skip from sheet top (ignored when section_start_row > 0)
    section_start_row: int = 0, # 1-based absolute row to begin reading (0 = use skip_rows)
    section_end_row: int = 0,   # 1-based absolute row to stop reading (0 = read to end)
) -> List[Dict[str, Any]]:
    """
    Read the sheet and return one dict per non-empty label row with:
      - label      : str   — text from the first non-empty cell in the row
      - label_col  : int   — 1-based column index where the label was found
      - value      : float | None
      - bold       : bool
      - italic     : bool
      - indent     : int   — alignment indent level (0 = leftmost)
      - row_index  : int   — 1-based sheet row number

    Row range: when section_start_row > 0, only rows in [section_start_row,
    section_end_row] are processed. This allows multi-statement sheets to be
    split correctly. Falls back to skip_rows when section_start_row is not set.

    Scaling is applied: values are normalised to actual dollars.
    Rows with no label text are skipped.
    """
    scale = _parse_scale(source_scaling)

    # Resolve effective row bounds
    if section_start_row > 0:
        start_row = section_start_row
    else:
        start_row = skip_rows + 1  # 1-based: skip_rows=3 means start at row 4

    end_row: Optional[int] = section_end_row if section_end_row > 0 else None

    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
    ws = wb[sheet_name]

    rows: List[Dict[str, Any]] = []

    for row_num, row in enumerate(ws.iter_rows(), start=1):
        if row_num < start_row:
            continue
        if end_row is not None and row_num > end_row:
            break

        # Find label: first non-empty cell in the row
        label: Optional[str] = None
        label_col: Optional[int] = None
        label_cell = None
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                label = str(cell.value).strip()
                label_col = cell.column
                label_cell = cell
                break

        if label is None:
            continue  # blank row

        # Value from the target column
        value_cell = ws.cell(row=row_num, column=column_index)
        raw_val = _cell_value(value_cell)
        raw_str = str(raw_val).strip() if raw_val is not None else ""

        font = _effective_font(label_cell)
        is_bold = bool(font.bold)
        indent = _indent_level(label_cell)

        if raw_val is None or raw_str == "":
            # Genuinely empty cell — title/header row, skip it
            continue
        elif raw_str in ("-", "—", "–"):
            # Dash placeholder — could be a genuine zero OR a section header.
            # Skip non-bold, low-indent rows with a dash value: these are almost
            # always section headers that happen to have a dash in the value column.
            if not is_bold and indent <= 1:
                continue
            value: float = 0.0
        else:
            try:
                value = float(raw_str.replace(",", "").replace("(", "-").replace(")", "")) * scale
            except (ValueError, TypeError):
                # Non-numeric (e.g. "N/A", text) — skip
                continue

        rows.append({
            "label": label,
            "label_col": label_col,
            "value": value,
            "bold": is_bold,
            "italic": bool(font.italic),
            "indent": indent,
            "row_index": row_num,
        })

    wb.close()
    return rows


def rows_to_csv_with_metadata(rows: List[Dict[str, Any]]) -> str:
    """
    Serialise the output of extract_rows_with_metadata to a CSV string.

    Columns: row_index, label, value, bold, italic, indent
    """
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=["row_index", "label", "value", "bold", "italic", "indent"],
        extrasaction="ignore",
        lineterminator="\n",
    )
    writer.writeheader()
    for r in rows:
        writer.writerow({
            "row_index": r["row_index"],
            "label": r["label"],
            "value": r["value"],
            "bold": r["bold"],
            "italic": r["italic"],
            "indent": r["indent"],
        })
    return output.getvalue()


# ── internal ──────────────────────────────────────────────────────────────────

def _parse_scale(source_scaling: str) -> float:
    s = source_scaling.lower()
    if "thousand" in s:
        return 1_000.0
    if "million" in s:
        return 1_000_000.0
    return 1.0
