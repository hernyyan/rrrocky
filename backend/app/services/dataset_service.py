"""
DatasetService — append Layer 1 extraction data to a company's Excel dataset.

Deep interface: callers provide company name, period, and raw layer1_results dict;
all Excel manipulation, fuzzy label matching via Claude, and file I/O are hidden
behind a single append_period() call.
"""
import json
import os
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles import numbers as xl_numbers
from openpyxl.utils import get_column_letter

from app.config import COMPANY_DATASETS_DIR
from app.services.claude_service import ClaudeService, get_claude_service

SHEET_NAME = "Financials"
IS_HEADER = "═══ INCOME STATEMENT ═══"
BS_HEADER = "═══ BALANCE SHEET ═══"


class DatasetAppendResult:
    __slots__ = ("company_name", "year", "period", "is_items_count", "bs_items_count", "label_changes")

    def __init__(self, company_name, year, period, is_items_count, bs_items_count, label_changes):
        self.company_name = company_name
        self.year = year
        self.period = period
        self.is_items_count = is_items_count
        self.bs_items_count = bs_items_count
        self.label_changes = label_changes

    def to_dict(self) -> dict:
        return {
            "success": True,
            "company_name": self.company_name,
            "year": self.year,
            "period": self.period,
            "is_items_count": self.is_items_count,
            "bs_items_count": self.bs_items_count,
            "label_changes": self.label_changes,
        }


class DatasetService:
    def __init__(self, claude: ClaudeService, datasets_dir: Path = COMPANY_DATASETS_DIR):
        self._claude = claude
        self._datasets_dir = datasets_dir

    # ── Public interface ────────────────────────────────────────────────────────

    def append_period(
        self,
        company_name: str,
        reporting_period: str,
        layer1_results: dict,
    ) -> DatasetAppendResult:
        """
        Append a new reporting period to the company's accumulating Excel dataset.

        Raises ValueError for invalid inputs (bad period, no items, duplicate period).
        All file and Claude side effects are contained here.
        """
        company_name = company_name.strip()
        if not company_name:
            raise ValueError("Company name is required.")

        period = reporting_period.strip()
        year = _extract_year(period)
        if not year:
            raise ValueError(f"Could not extract year from reporting period: '{period}'")

        is_items, bs_items = _split_layer1_results(layer1_results)
        if not is_items and not bs_items:
            raise ValueError("No line items to append.")

        company_dir = self._datasets_dir / _safe_dirname(company_name)
        company_dir.mkdir(parents=True, exist_ok=True)

        xlsx_path = company_dir / f"{year}.xlsx"
        changelog_path = company_dir / f"{year}_changelog.jsonl"

        wb, ws = _load_or_create_workbook(xlsx_path)

        existing_periods = _get_existing_periods(ws)
        if period in existing_periods:
            raise ValueError(f"Period '{period}' already exists in the dataset for {company_name} {year}.")

        existing_is_labels, existing_bs_labels = _get_existing_labels(ws)

        is_mapping, bs_mapping, changelog_entries = self._run_fuzzy_matching(
            existing_is_labels, existing_bs_labels, is_items, bs_items, period
        )

        _append_to_worksheet(ws, period, is_items, bs_items, is_mapping, bs_mapping)
        _apply_formatting(ws)

        wb.save(str(xlsx_path))

        if changelog_entries:
            with open(changelog_path, "a", encoding="utf-8") as f:
                for entry in changelog_entries:
                    f.write(json.dumps(entry) + "\n")

        return DatasetAppendResult(
            company_name=company_name,
            year=year,
            period=period,
            is_items_count=len(is_items),
            bs_items_count=len(bs_items),
            label_changes=len(changelog_entries),
        )

    # ── Fuzzy matching ──────────────────────────────────────────────────────────

    def _run_fuzzy_matching(
        self,
        existing_is_labels: list,
        existing_bs_labels: list,
        is_items: dict,
        bs_items: dict,
        period: str,
    ) -> tuple[dict, dict, list]:
        """Run Claude fuzzy matching for IS and BS labels. Returns (is_mapping, bs_mapping, changelog_entries)."""
        is_mapping: dict = {}
        bs_mapping: dict = {}
        changelog_entries: list = []

        if existing_is_labels and is_items:
            is_mapping, is_changes = self._fuzzy_match_labels(
                existing_is_labels, list(is_items.keys()), "income_statement", period
            )
            changelog_entries.extend(is_changes)

        if existing_bs_labels and bs_items:
            bs_mapping, bs_changes = self._fuzzy_match_labels(
                existing_bs_labels, list(bs_items.keys()), "balance_sheet", period
            )
            changelog_entries.extend(bs_changes)

        return is_mapping, bs_mapping, changelog_entries

    def _fuzzy_match_labels(
        self,
        existing_labels: list,
        incoming_labels: list,
        statement_type: str,
        period: str,
    ) -> tuple[dict, list]:
        """
        Ask Claude Sonnet to match incoming labels to existing labels.
        Returns (mapping, changelog_entries). Fails gracefully — returns empty on error.
        """
        if not existing_labels or not incoming_labels:
            return {}, []

        model = os.getenv("LAYER1_MODEL", "claude-sonnet-4-6")
        prompt = _build_fuzzy_match_prompt(existing_labels, incoming_labels)

        try:
            response_text = self._claude.call_claude_raw(prompt, model, max_tokens=4096)
        except Exception as e:
            print(f"Fuzzy match Claude call failed: {e}. Treating all labels as new.")
            return {}, []

        try:
            raw = self._claude.parse_json_response(response_text)
        except Exception as e:
            print(f"Fuzzy match JSON parse failed: {e}. Treating all labels as new.")
            return {}, []

        mapping: dict = {}
        changelog_entries: list = []

        for match in raw.get("matches", []):
            incoming = match.get("incoming", "")
            existing = match.get("existing", "")
            if incoming and existing and incoming != existing:
                mapping[incoming] = existing
                changelog_entries.append({
                    "timestamp": datetime.utcnow().isoformat(),
                    "statement_type": statement_type,
                    "period": period,
                    "current_label": incoming,
                    "previous_label": existing,
                    "action": "renamed",
                    "all_previous_names": [],  # populated as changelog accumulates
                })

        return mapping, changelog_entries


# ── Module-level helpers (pure functions, no side effects) ──────────────────────

def _split_layer1_results(layer1_results: dict) -> tuple[dict, dict]:
    """Extract IS and BS line items from the Layer 1 results dict."""
    is_items: dict = {}
    bs_items: dict = {}
    for stmt_type, result_data in layer1_results.items():
        items = result_data.get("lineItems", {})
        if "income" in stmt_type.lower():
            is_items = items
        elif "balance" in stmt_type.lower():
            bs_items = items
    return is_items, bs_items


def _load_or_create_workbook(xlsx_path: Path):
    """Load existing workbook or create a new one."""
    if xlsx_path.exists():
        wb = load_workbook(str(xlsx_path))
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
    return wb, ws


def _get_existing_periods(ws) -> list:
    """Read all period column headers from row 1, starting at column B."""
    periods = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            periods.append(str(val))
    return periods


def _get_existing_labels(ws) -> tuple[list, list]:
    """
    Read existing line item labels from column A.
    Returns (is_labels, bs_labels) split by the section headers.
    """
    is_labels: list = []
    bs_labels: list = []
    current_section = None

    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val == IS_HEADER:
            current_section = "is"
            continue
        elif val == BS_HEADER:
            current_section = "bs"
            continue
        elif val is None or str(val).strip() == "":
            continue
        elif row == 1:
            continue  # header row

        if current_section == "is":
            is_labels.append(str(val))
        elif current_section == "bs":
            bs_labels.append(str(val))

    return is_labels, bs_labels


def _append_to_worksheet(ws, period, is_items, bs_items, is_mapping, bs_mapping):
    """
    Add a new column for this period with all line items.
    Handles matching existing rows via mapping and inserting new rows.
    """
    if ws.max_row <= 1 and ws.cell(row=1, column=1).value is None:
        _build_new_sheet(ws, period, is_items, bs_items)
        return

    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value=period)

    label_row_map = {}
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and str(val).strip() and val not in (IS_HEADER, BS_HEADER):
            label_row_map[str(val)] = row

    def find_or_create_row(label, mapping, section_header):
        mapped_label = mapping.get(label, label)
        if mapped_label in label_row_map:
            row = label_row_map[mapped_label]
            if mapped_label != label:
                ws.cell(row=row, column=1, value=label)
                label_row_map[label] = row
                del label_row_map[mapped_label]
            return row
        insert_row = _find_insert_position(ws, section_header)
        ws.insert_rows(insert_row)
        ws.cell(row=insert_row, column=1, value=label)
        label_row_map[label] = insert_row
        return insert_row

    for label, value in is_items.items():
        row = find_or_create_row(label, is_mapping, IS_HEADER)
        ws.cell(row=row, column=new_col, value=value)

    for label, value in bs_items.items():
        row = find_or_create_row(label, bs_mapping, BS_HEADER)
        ws.cell(row=row, column=new_col, value=value)


def _build_new_sheet(ws, period, is_items, bs_items):
    """Build the worksheet from scratch for the first period."""
    row = 1
    ws.cell(row=row, column=1, value="Line Item")
    ws.cell(row=row, column=2, value=period)

    row += 1
    ws.cell(row=row, column=1, value=IS_HEADER)

    for label, value in is_items.items():
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)

    row += 1  # blank separator

    row += 1
    ws.cell(row=row, column=1, value=BS_HEADER)

    for label, value in bs_items.items():
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)


def _find_insert_position(ws, section_header) -> int:
    """Find the row to insert a new label at the end of a section."""
    section_start = None
    for row in range(1, ws.max_row + 2):
        val = ws.cell(row=row, column=1).value if row <= ws.max_row else None
        if val == section_header:
            section_start = row
        elif section_start is not None and (
            val == BS_HEADER or val == IS_HEADER or row > ws.max_row
        ):
            for r in range(row - 1, section_start, -1):
                cell_val = ws.cell(row=r, column=1).value
                if cell_val and str(cell_val).strip():
                    return r + 1
            return section_start + 1
    return ws.max_row + 1


def _apply_formatting(ws):
    """Apply basic styling to the worksheet."""
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")
    left_align = Alignment(horizontal="left")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    number_format = "#,##0.00_);(#,##0.00)"

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.alignment = center_align

    ws.column_dimensions["A"].width = 40
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    for row in range(2, ws.max_row + 1):
        label_cell = ws.cell(row=row, column=1)
        val = label_cell.value
        if val in (IS_HEADER, BS_HEADER):
            label_cell.font = bold_font
            label_cell.fill = gray_fill
            label_cell.alignment = left_align
            for col in range(2, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = gray_fill
        else:
            label_cell.alignment = left_align
            for col in range(2, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.number_format = number_format
                    cell.alignment = right_align


def _build_fuzzy_match_prompt(existing_labels: list, incoming_labels: list) -> str:
    return f"""You are matching financial statement line item labels between two reporting periods for the same company.

## Existing labels (already in the dataset):
{json.dumps(existing_labels)}

## Incoming labels (new period being added):
{json.dumps(incoming_labels)}

## Task

For each incoming label, determine if it is the SAME line item as one of the existing labels, just with a different name. Consider:
- Abbreviation differences: "SG&A" = "Selling, General & Administrative"
- Word order changes: "Total Cost of Revenue" = "Cost of Revenue, Total"
- Minor wording changes: "Cost of Sales" = "Cost of Goods Sold"
- Punctuation/formatting differences: "Acc. Depreciation" = "Accumulated Depreciation"

Only match labels that clearly refer to the same financial concept. If there is ANY uncertainty, do NOT match — leave the incoming label as new.

Do NOT match labels just because they are subtotals or totals of different things (e.g., "Total Current Assets" and "Total Assets" are NOT the same).

## Output

Return a JSON object with exactly one key:

```json
{{
  "matches": [
    {{
      "incoming": "Cost of Goods Sold",
      "existing": "Cost of Sales",
      "confidence": "high"
    }}
  ]
}}
```

Only include matches where confidence is "high". If an incoming label has no match among existing labels, do NOT include it in the output — it will be treated as a new row. Return an empty matches array if there are no confident matches.
"""


def _extract_year(period: str) -> str | None:
    """Extract 4-digit year from a period string like 'July 2025'."""
    match = re.search(r"20\d{2}", period)
    return match.group(0) if match else None


def _safe_dirname(name: str) -> str:
    """Sanitize company name for use as directory name."""
    return re.sub(r'[^\w\s\-]', '', name).strip().replace(' ', '_')


# ── Factory ─────────────────────────────────────────────────────────────────────

_instance: DatasetService | None = None


def get_dataset_service() -> DatasetService:
    global _instance
    if _instance is None:
        _instance = DatasetService(claude=get_claude_service())
    return _instance
