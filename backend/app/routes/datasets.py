"""
POST /datasets/append — Append Layer 1 raw extraction data to the company's
accumulating Excel dataset. Runs fuzzy label matching via Claude Sonnet.
"""
import json
import os
import re
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles import numbers as xl_numbers

from app.config import COMPANY_DATASETS_DIR
from app.models.schemas import DatasetAppendRequest
from app.services.claude_service import get_claude_service

router = APIRouter()

SHEET_NAME = "Financials"
IS_HEADER = "═══ INCOME STATEMENT ═══"
BS_HEADER = "═══ BALANCE SHEET ═══"


@router.post("/datasets/append")
def append_to_dataset(request: DatasetAppendRequest):
    """
    Append Layer 1 results for a reporting period to the company's
    yearly Excel dataset. Runs fuzzy matching on line item labels.
    """
    company_name = request.company_name.strip()
    if not company_name:
        raise HTTPException(status_code=400, detail="Company name is required.")

    period = request.reporting_period.strip()
    year = _extract_year(period)
    if not year:
        raise HTTPException(
            status_code=400,
            detail=f"Could not extract year from reporting period: '{period}'",
        )

    # Prepare directory
    company_dir = COMPANY_DATASETS_DIR / _safe_dirname(company_name)
    company_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = company_dir / f"{year}.xlsx"
    changelog_path = company_dir / f"{year}_changelog.jsonl"

    # Load or create workbook
    if xlsx_path.exists():
        wb = load_workbook(str(xlsx_path))
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

    # Check for duplicate period
    existing_periods = _get_existing_periods(ws)
    if period in existing_periods:
        raise HTTPException(
            status_code=409,
            detail=f"Period '{period}' already exists in the dataset for {company_name} {year}.",
        )

    # Extract incoming line items
    is_items = {}
    bs_items = {}
    for stmt_type, result_data in request.layer1_results.items():
        items = result_data.get("lineItems", {})
        if "income" in stmt_type.lower():
            is_items = items
        elif "balance" in stmt_type.lower():
            bs_items = items

    if not is_items and not bs_items:
        raise HTTPException(status_code=400, detail="No line items to append.")

    # Get existing labels from the worksheet
    existing_is_labels, existing_bs_labels = _get_existing_labels(ws)

    # Run fuzzy matching if there are existing labels to match against
    is_mapping = {}
    bs_mapping = {}
    changelog_entries = []

    if existing_is_labels and is_items:
        is_mapping, is_changes = _run_fuzzy_match(
            existing_is_labels, list(is_items.keys()), "income_statement", period
        )
        changelog_entries.extend(is_changes)

    if existing_bs_labels and bs_items:
        bs_mapping, bs_changes = _run_fuzzy_match(
            existing_bs_labels, list(bs_items.keys()), "balance_sheet", period
        )
        changelog_entries.extend(bs_changes)

    # Apply mappings and append data
    _append_to_worksheet(ws, period, is_items, bs_items, is_mapping, bs_mapping)

    # Apply formatting
    _apply_formatting(ws)

    # Save workbook
    wb.save(str(xlsx_path))

    # Append changelog entries
    if changelog_entries:
        with open(changelog_path, "a", encoding="utf-8") as f:
            for entry in changelog_entries:
                f.write(json.dumps(entry) + "\n")

    return {
        "success": True,
        "company_name": company_name,
        "year": year,
        "period": period,
        "is_items_count": len(is_items),
        "bs_items_count": len(bs_items),
        "label_changes": len(changelog_entries),
    }


# ─── Fuzzy matching ────────────────────────────────────────────────────────────

def _run_fuzzy_match(
    existing_labels: list,
    incoming_labels: list,
    statement_type: str,
    period: str,
) -> tuple:
    """
    Use Claude Sonnet to match incoming labels to existing labels.

    Returns:
        - mapping: dict of { incoming_label: existing_label } for matches.
          Incoming labels not in this dict are treated as new rows.
        - changelog_entries: list of dicts for labels that were renamed.
    """
    if not existing_labels or not incoming_labels:
        return {}, []

    claude = get_claude_service()
    model = os.getenv("LAYER1_MODEL", "claude-sonnet-4-6")

    prompt = f"""You are matching financial statement line item labels between two reporting periods for the same company.

## Existing labels (already in the dataset):
{json.dumps(existing_labels)}

## Incoming labels (new period being added):
{json.dumps(incoming_labels)}

## Task

For each incoming label, determine if it is the SAME line item as one of the existing labels, just with a different name. Consider:
- Abbreviation differences: "SG&A" = "Selling, General & Administrative"
- Word order changes: "Total Cost of Revenue" = "Cost of Revenue, Total"
- Minor wording changes: "Cost of Sales" = "Cost of Goods Sold"
- Punctuation/formatting differences: "Acc. Depreciation" = "Accumulated Depreciation"

Only match labels that clearly refer to the same financial concept. If there is ANY uncertainty, do NOT match — leave the incoming label as new.

Do NOT match labels just because they are subtotals or totals of different things (e.g., "Total Current Assets" and "Total Assets" are NOT the same).

## Output

Return a JSON object with exactly one key:

```json
{{
  "matches": [
    {{
      "incoming": "Cost of Goods Sold",
      "existing": "Cost of Sales",
      "confidence": "high"
    }}
  ]
}}
```

Only include matches where confidence is "high". If an incoming label has no match among existing labels, do NOT include it in the output — it will be treated as a new row. Return an empty matches array if there are no confident matches.
"""

    try:
        response_text = claude.call_claude_raw(prompt, model, max_tokens=4096)
    except Exception as e:
        print(f"Fuzzy match Claude call failed: {e}. Treating all labels as new.")
        return {}, []

    try:
        raw = claude.parse_json_response(response_text)
    except Exception as e:
        print(f"Fuzzy match JSON parse failed: {e}. Treating all labels as new.")
        return {}, []

    matches = raw.get("matches", [])

    mapping = {}
    changelog_entries = []

    for match in matches:
        incoming = match.get("incoming", "")
        existing = match.get("existing", "")
        if incoming and existing and incoming != existing:
            mapping[incoming] = existing
            changelog_entries.append({
                "timestamp": datetime.utcnow().isoformat(),
                "statement_type": statement_type,
                "period": period,
                "current_label": incoming,
                "previous_label": existing,
                "action": "renamed",
                "all_previous_names": _get_label_history(existing, statement_type),
            })

    return mapping, changelog_entries


# ─── Worksheet operations ──────────────────────────────────────────────────────

def _get_existing_periods(ws) -> list:
    """Read all period column headers from row 1, starting at column B."""
    periods = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            periods.append(str(val))
    return periods


def _get_existing_labels(ws) -> tuple:
    """
    Read existing line item labels from column A.
    Returns (is_labels, bs_labels) split by the section headers.
    """
    is_labels = []
    bs_labels = []
    current_section = None

    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val == IS_HEADER:
            current_section = "is"
            continue
        elif val == BS_HEADER:
            current_section = "bs"
            continue
        elif val is None or str(val).strip() == "":
            continue
        elif row == 1:
            continue  # header row

        if current_section == "is":
            is_labels.append(str(val))
        elif current_section == "bs":
            bs_labels.append(str(val))

    return is_labels, bs_labels


def _append_to_worksheet(ws, period, is_items, bs_items, is_mapping, bs_mapping):
    """
    Add a new column for this period with all line items.
    Handles matching existing rows via mapping and inserting new rows.
    """
    # If sheet is empty, build it from scratch
    if ws.max_row <= 1 and ws.cell(row=1, column=1).value is None:
        _build_new_sheet(ws, period, is_items, bs_items)
        return

    # Determine the new column index
    new_col = ws.max_column + 1

    # Write period header
    ws.cell(row=1, column=new_col, value=period)

    # Build a label → row index map from existing data
    label_row_map = {}
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and str(val).strip() and val not in (IS_HEADER, BS_HEADER):
            label_row_map[str(val)] = row

    def find_or_create_row(label, mapping, section_header):
        mapped_label = mapping.get(label, label)

        if mapped_label in label_row_map:
            row = label_row_map[mapped_label]
            if mapped_label != label:
                ws.cell(row=row, column=1, value=label)
                label_row_map[label] = row
                del label_row_map[mapped_label]
            return row

        # New label — insert at end of section
        insert_row = _find_insert_position(ws, section_header)
        ws.insert_rows(insert_row)
        ws.cell(row=insert_row, column=1, value=label)
        label_row_map[label] = insert_row
        return insert_row

    for label, value in is_items.items():
        row = find_or_create_row(label, is_mapping, IS_HEADER)
        ws.cell(row=row, column=new_col, value=value)

    for label, value in bs_items.items():
        row = find_or_create_row(label, bs_mapping, BS_HEADER)
        ws.cell(row=row, column=new_col, value=value)


def _build_new_sheet(ws, period, is_items, bs_items):
    """Build the worksheet from scratch for the first period."""
    row = 1
    ws.cell(row=row, column=1, value="Line Item")
    ws.cell(row=row, column=2, value=period)

    row += 1
    ws.cell(row=row, column=1, value=IS_HEADER)

    for label, value in is_items.items():
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)

    row += 1  # blank separator row

    row += 1
    ws.cell(row=row, column=1, value=BS_HEADER)

    for label, value in bs_items.items():
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)


def _find_insert_position(ws, section_header) -> int:
    """Find the row to insert a new label at the end of a section."""
    section_start = None
    for row in range(1, ws.max_row + 2):
        val = ws.cell(row=row, column=1).value if row <= ws.max_row else None
        if val == section_header:
            section_start = row
        elif section_start is not None and (
            val == BS_HEADER or val == IS_HEADER or row > ws.max_row
        ):
            # Walk backwards to find the last non-empty row in this section
            for r in range(row - 1, section_start, -1):
                cell_val = ws.cell(row=r, column=1).value
                if cell_val and str(cell_val).strip():
                    return r + 1
            return section_start + 1

    return ws.max_row + 1


def _apply_formatting(ws):
    """Apply basic styling to the worksheet."""
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")
    left_align = Alignment(horizontal="left")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    number_format = "#,##0.00_);(#,##0.00)"

    # Row 1: bold + center for all cells
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.alignment = center_align

    # Column A: set width
    ws.column_dimensions["A"].width = 40

    # Period columns: set width
    from openpyxl.utils import get_column_letter
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Iterate rows to style section headers and values
    for row in range(2, ws.max_row + 1):
        label_cell = ws.cell(row=row, column=1)
        val = label_cell.value

        if val in (IS_HEADER, BS_HEADER):
            label_cell.font = bold_font
            label_cell.fill = gray_fill
            label_cell.alignment = left_align
            # Apply gray fill across all columns for section header rows
            for col in range(2, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = gray_fill
        else:
            label_cell.alignment = left_align
            # Format value columns
            for col in range(2, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.number_format = number_format
                    cell.alignment = right_align


# ─── Utilities ─────────────────────────────────────────────────────────────────

def _extract_year(period: str):
    """Extract 4-digit year from a period string like 'July 2025'."""
    match = re.search(r"20\d{2}", period)
    return match.group(0) if match else None


def _safe_dirname(name: str) -> str:
    """Sanitize company name for use as directory name."""
    return re.sub(r'[^\w\s\-]', '', name).strip().replace(' ', '_')


def _get_label_history(label: str, statement_type: str) -> list:
    """
    Look up the full rename history for a label from the changelog.
    For now, returns empty list — will be populated as changelog accumulates.
    """
    # TODO: Read changelog and build full history chain
    return []
